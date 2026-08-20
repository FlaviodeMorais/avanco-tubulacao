# -*- coding: utf-8 -*-
"""Gera a planilha de retomada (ControlTub-Retomada-Out2026.xlsx) a partir dos
dois arquivos que o sistema ControlTub sempre exporta nesse mesmo padrão:

  - Mapa de Spools (aba "SGS", cabeçalho na linha 9)
  - Mapa de Juntas (aba "MAPA_JUNTA", cabeçalho na linha 8)

Um registro é "remanescente" quando não tem data de "Liberado END" — o mesmo
marco de conclusão que o painel web (ControlTub-Dashboard.V5.5.html) já usa:
campo "libcdat" para spools, campo "dt_lib_END" para juntas.

Uso:
    python3 retomada/build_retomada.py --spools SGS.xlsx --juntas MAPA_JUNTA.xlsx

Para atualizar com uma exportação mais recente do ControlTub (mesmo padrão),
rode de novo apontando para os arquivos novos — não precisa editar nada aqui.
"""
import argparse
import calendar
import datetime
from collections import defaultdict

import openpyxl
from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

COLUNAS_SPOOLS = [
    'unid', 'subunid', 'isometrico + - + spool', 'fabricante', 'tp_linha', 'linha',
    'fluido', 'revisao', 'mat', 'diametro', 'dipol', 'esp', 'hold', 'espec', 'comp',
    'area', 'peso', 'copi', 'TIPO', 'tjs', 'tjsp', 'Juntas_Campo', 'sth', 'sop',
    'subsop', 'dtlibfab', 'progfab', 'dtcort', 'ajudat', 'soldat', 'vsdat', 'libdat',
    'dtdim', 'dtpint', 'Aju_P', 'Sold_P', 'Apro_P', 'ROMANEIO', 'DTEMBARQUE', 'sger',
    'dtprogmon', 'dtpremon', 'dtposmon', 'ajucdat', 'solcdat', 'vscdat', 'libcdat',
    'dtdic', 'pintacab', 'Aju_C', 'Sold_C', 'Aprov_C', 'isol', 'sthdat', 'lavdat',
    'condat', 'isdat', 'supdat', 'acedat', 'docdat', 'sgermon', 'Cor', 'TR_SP',
    'Setor', 'dtlibmon', 'PROGMON', 'DT_RET_PIN',
]
COLUNAS_JUNTAS = [
    'unid', 'subunid', 'sth', 'isometrico', 'spool', 'junta', 'linha', 'espec',
    'fluido', 'P_C', 'tipo', 'diam', 'pol', 'esp', 'mat', 'NI', 'pn', 'Tub_Con',
    'TT', 'DU', 'OBS', 'progfab', 'dtcort', 'dtprogmon', 'dtpremon', 'dtaju',
    'asraiz', 'asench', 'ProcRaiz', 'ProcEnch', 'IEIS', 'dtsold', 'LTraiz',
    'LTench', 'dtsoldr1', 'dtsoldr2', 'corrida1', 'corrida2', 'dtVS', 'LP_Acab',
    'LP_antes_TT', 'dtTT', 'dtDU', 'RX_US', 'US1', 'dtUS', 'US2', 'dtUS2', 'US3',
    'dtUS3', 'US4', 'dtUS4', 'dt_lib_END', 'dtDI_PIPE', 'dtDI_CAMPO',
    'HT_NUMBER1', 'HT_NUMBER2', 'sger', 'DTEMBARQUE', 'CBES', 'laudo DU',
    'FABRICANTE',
]

AZUL, AZUL_CLARO, BRANCO = 'FF1F4E78', 'FFDCE6F1', 'FFFFFFFF'
FONTE_TITULO = Font(name='Calibri', size=14, bold=True, color=BRANCO)
FONTE_SUBTITULO = Font(name='Calibri', size=10, italic=True, color='FF595959')
FONTE_CABECALHO = Font(name='Calibri', size=10, bold=True, color=BRANCO)
FONTE_ROTULO = Font(name='Calibri', size=10, bold=True)
FONTE_NORMAL = Font(name='Calibri', size=10)
PREENCH_TITULO = PatternFill('solid', fgColor=AZUL)
PREENCH_CABECALHO = PatternFill('solid', fgColor=AZUL)
PREENCH_ROTULO = PatternFill('solid', fgColor=AZUL_CLARO)
BORDA_FINA = Border(*(Side(style='thin', color='FFBFBFBF'),) * 4)
FMT_TON, FMT_PCT, FMT_INT, FMT_DATA = '#,##0.000', '0.0%', '#,##0', 'dd/mm/yyyy'
MESES_ABREV = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']


def cabecalho(ws, linha):
    valores = next(ws.iter_rows(min_row=linha, max_row=linha, values_only=True))
    return {str(v).strip(): i for i, v in enumerate(valores) if v is not None}


def carregar_linhas(caminho, aba, linha_cabecalho, colunas_obrigatorias, filtro):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    if aba not in wb.sheetnames:
        raise SystemExit('Aba "%s" não encontrada em %s (abas: %s)' % (aba, caminho, wb.sheetnames))
    ws = wb[aba]
    H = cabecalho(ws, linha_cabecalho)
    faltando = [c for c in colunas_obrigatorias if c not in H]
    if faltando:
        raise SystemExit('Colunas ausentes em %s (aba %s): %s' % (caminho, aba, faltando))
    linhas = []
    for tupla in ws.iter_rows(min_row=linha_cabecalho + 1, values_only=True):
        row = {nome: tupla[idx] for nome, idx in H.items()}
        if filtro(row):
            linhas.append(row)
    wb.close()
    return linhas


def agregados(validas_sp, chave):
    agg = defaultdict(lambda: {'peso_total': 0.0, 'peso_remanescente': 0.0,
                               'spools_total': 0, 'spools_remanescente': 0, 'linhas': set()})
    for r in validas_sp:
        a = agg[r.get(chave) or '(sem %s)' % chave]
        a['peso_total'] += r['peso']
        a['spools_total'] += 1
        if r.get('linha'):
            a['linhas'].add(r['linha'])
        if not r.get('libcdat'):
            a['peso_remanescente'] += r['peso']
            a['spools_remanescente'] += 1
    return {k: {**a, 'linhas': len(a['linhas'])} for k, a in agg.items()}


def titulo_pagina(ws, texto, subtexto, largura_colunas):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=largura_colunas)
    c = ws.cell(1, 1, texto)
    c.font, c.fill = FONTE_TITULO, PREENCH_TITULO
    c.alignment = Alignment(vertical='center', horizontal='left', indent=1)
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=largura_colunas)
    c = ws.cell(2, 1, subtexto)
    c.font = FONTE_SUBTITULO
    c.alignment = Alignment(vertical='center', horizontal='left', indent=1)


def escrever_tabela_bruta(ws, colunas, linhas, linha_inicial, nome_tabela):
    for j, nome in enumerate(colunas, start=1):
        c = ws.cell(linha_inicial, j, nome)
        c.font, c.fill = FONTE_CABECALHO, PREENCH_CABECALHO
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[linha_inicial].height = 30
    for i, row in enumerate(linhas, start=linha_inicial + 1):
        for j, nome in enumerate(colunas, start=1):
            valor = row.get(nome)
            c = ws.cell(i, j, valor)
            if isinstance(valor, (datetime.date, datetime.datetime)):
                c.number_format = FMT_DATA
    ultima_linha = linha_inicial + len(linhas)
    ref = 'A%d:%s%d' % (linha_inicial, get_column_letter(len(colunas)), ultima_linha)
    tabela = Table(displayName=nome_tabela, ref=ref)
    tabela.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    ws.add_table(tabela)
    ws.freeze_panes = ws.cell(linha_inicial + 1, 1).coordinate
    for j, nome in enumerate(colunas, start=1):
        largura = 26 if ('isometrico' in nome or nome in
                         ('linha', 'espec', 'sger', 'sgermon', 'OBS', 'sop', 'subsop', 'sth')) else 13
        ws.column_dimensions[get_column_letter(j)].width = largura


def cartao(ws, linha, col, rotulo, valor, fmt):
    c1 = ws.cell(linha, col, rotulo)
    c1.font, c1.fill, c1.border = FONTE_ROTULO, PREENCH_ROTULO, BORDA_FINA
    c1.alignment = Alignment(vertical='center', indent=1)
    c2 = ws.cell(linha, col + 1, valor)
    c2.font = Font(name='Calibri', size=12, bold=True, color=AZUL)
    c2.number_format, c2.border = fmt, BORDA_FINA
    c2.alignment = Alignment(vertical='center', horizontal='right', indent=1)
    ws.row_dimensions[linha].height = 20


def montar(spools_path, juntas_path, saida, data_inicio, prazo_meses):
    validas_sp = carregar_linhas(
        spools_path, 'SGS', 9, ['unid', 'peso', 'libcdat'],
        lambda r: r.get('unid') and r.get('peso') is not None)
    remanescentes_sp = [r for r in validas_sp if not r.get('libcdat')]
    concluidos_sp_n = len(validas_sp) - len(remanescentes_sp)

    validas_jt = carregar_linhas(
        juntas_path, 'MAPA_JUNTA', 8, ['unid', 'junta', 'dt_lib_END'],
        lambda r: r.get('unid') and r.get('junta'))
    remanescentes_jt = [r for r in validas_jt if not r.get('dt_lib_END')]
    concluidas_jt_n = len(validas_jt) - len(remanescentes_jt)

    peso_total = sum(r['peso'] for r in validas_sp) / 1000
    peso_remanescente = round(sum(r['peso'] for r in remanescentes_sp) / 1000, 3)
    peso_concluido = round(peso_total - peso_remanescente, 3)
    pct_remanescente = peso_remanescente / peso_total if peso_total else 0

    por_unidade = agregados(validas_sp, 'unid')
    por_sop = agregados(validas_sp, 'sop')

    mes_fim = (data_inicio.month - 1 + prazo_meses - 1) % 12 + 1
    ano_fim = data_inicio.year + (data_inicio.month - 1 + prazo_meses - 1) // 12
    data_fim = datetime.date(ano_fim, mes_fim, calendar.monthrange(ano_fim, mes_fim)[1])

    wb = openpyxl.Workbook()

    # --------------------------------------------------------- Leia-me ----
    ws = wb.active
    ws.title = 'Leia-me'
    titulo_pagina(ws, 'ControlTub — Retomada de Fabricação e Montagem',
                  'Fonte: %s + %s' % (spools_path.split('/')[-1], juntas_path.split('/')[-1]), 6)
    linhas_texto = [
        (4, 'O que é este arquivo', True),
        (5, 'Separa, dos dois arquivos originais do ControlTub (mesmo padrão sempre exportado), o que já foi', False),
        (6, 'concluído do que ainda falta fazer — o "remanescente" — para planejar a retomada da obra.', False),
        (8, 'Critério de remanescente (o que ainda falta)', True),
        (9, 'Um registro entra como remanescente quando não tem data de "Liberado END" — o mesmo marco de', False),
        (10, 'conclusão que o painel web (avanco-tubulacao.vercel.app) já usa: campo "libcdat" para spools', False),
        (11, '(aba SGS) e campo "dt_lib_END" para juntas (aba MAPA_JUNTA). Sem essa data, o item nunca foi', False),
        (12, 'liberado — está parado desde a paralisação do contrato anterior.', False),
        (14, 'Datas do novo contrato', True),
        (15, 'Início da retomada: %s  •  Prazo contratual: %d meses  •  Término previsto: %s'
         % (data_inicio.strftime('%d/%m/%Y'), prazo_meses, data_fim.strftime('%d/%m/%Y')), False),
        (17, 'Abas deste arquivo', True),
    ]
    for linha, texto, negrito in linhas_texto:
        c = ws.cell(linha, 1, texto)
        c.font = FONTE_ROTULO if negrito else FONTE_NORMAL
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=6)
    tabela_abas = [
        ('Resumo', 'Números-chave da retomada: escopo concluído x remanescente.'),
        ('Remanescentes_Spools', '%d spools sem Liberado END — %.3f ton, %.1f%% do peso total do projeto.'
         % (len(remanescentes_sp), peso_remanescente, pct_remanescente * 100)),
        ('Remanescentes_Juntas', '%d juntas sem Liberado END — de %d juntas cadastradas no total.'
         % (len(remanescentes_jt), len(validas_jt))),
        ('Acompanhamento_Fab_Montagem', 'Escopo novo por unidade/pacote SOP + cronograma mensal para preencher o avanço realizado.'),
    ]
    r0 = 18
    for j, cab in enumerate(('Aba', 'Conteúdo'), start=1):
        c = ws.cell(r0, j, cab)
        c.font, c.fill = FONTE_CABECALHO, PREENCH_CABECALHO
    for i, (aba, desc) in enumerate(tabela_abas, start=r0 + 1):
        ws.cell(i, 1, aba).font = FONTE_ROTULO
        ws.cell(i, 2, desc).font = FONTE_NORMAL
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        ws.row_dimensions[i].height = 28
        for j in (1, 2):
            ws.cell(i, j).alignment = Alignment(vertical='center', wrap_text=(j == 2))
    nota_linha = r0 + len(tabela_abas) + 2
    ws.cell(nota_linha, 1, 'Planilhas originais usadas (não alteradas):').font = FONTE_ROTULO
    ws.cell(nota_linha + 1, 1, '  • %s (aba SGS, %d spools válidos)' % (spools_path.split('/')[-1], len(validas_sp))).font = FONTE_NORMAL
    ws.cell(nota_linha + 2, 1, '  • %s (aba MAPA_JUNTA, %d juntas válidas)' % (juntas_path.split('/')[-1], len(validas_jt))).font = FONTE_NORMAL
    ws.column_dimensions['A'].width = 26
    for col in 'BCDEF':
        ws.column_dimensions[col].width = 16
    ws.sheet_view.showGridLines = False

    # ----------------------------------------------------------- Resumo ----
    ws = wb.create_sheet('Resumo')
    titulo_pagina(ws, 'Resumo da Retomada',
                  'Peso em toneladas — spools; juntas não têm peso próprio (peso é do spool)', 5)
    cartoes = [
        ('Peso total do projeto (ton)', peso_total, FMT_TON),
        ('Peso já concluído (Liberado END, ton)', peso_concluido, FMT_TON),
        ('Peso remanescente — escopo novo (ton)', peso_remanescente, FMT_TON),
        ('% remanescente do peso total', pct_remanescente, FMT_PCT),
        ('Spools remanescentes (qtd)', len(remanescentes_sp), FMT_INT),
        ('Spools concluídos (qtd)', concluidos_sp_n, FMT_INT),
        ('Juntas remanescentes (qtd)', len(remanescentes_jt), FMT_INT),
        ('Juntas concluídas (qtd)', concluidas_jt_n, FMT_INT),
        ('Data de início da retomada', data_inicio, FMT_DATA),
        ('Prazo contratual (meses)', prazo_meses, FMT_INT),
    ]
    for i, (rotulo, valor, fmt) in enumerate(cartoes, start=4):
        cartao(ws, i, 1, rotulo, valor, fmt)
    ws.column_dimensions['A'].width, ws.column_dimensions['B'].width, ws.column_dimensions['C'].width = 34, 16, 3
    ws.sheet_view.showGridLines = False
    ws.cell(3, 5, 'Concluído x Remanescente (ton)').font = FONTE_ROTULO
    ws.cell(4, 5, 'Concluído'); ws.cell(4, 6, peso_concluido)
    ws.cell(5, 5, 'Remanescente'); ws.cell(5, 6, peso_remanescente)
    grafico = PieChart()
    grafico.title = 'Peso do projeto (ton)'
    grafico.add_data(Reference(ws, min_col=6, min_row=4, max_row=5))
    grafico.set_categories(Reference(ws, min_col=5, min_row=4, max_row=5))
    grafico.height, grafico.width = 7, 11
    ws.add_chart(grafico, 'E7')

    # ------------------------------------------------ Remanescentes (raw) --
    ws = wb.create_sheet('Remanescentes_Spools')
    titulo_pagina(ws, 'Remanescentes — Spools',
                  'Sem "Liberado END" (libcdat) — %d de %d spools, %.3f ton | Fonte: %s, aba SGS'
                  % (len(remanescentes_sp), len(validas_sp), peso_remanescente, spools_path.split('/')[-1]),
                  len(COLUNAS_SPOOLS))
    escrever_tabela_bruta(ws, COLUNAS_SPOOLS, remanescentes_sp, 4, 'TabRemanescentesSpools')
    ws.sheet_view.showGridLines = False

    ws = wb.create_sheet('Remanescentes_Juntas')
    titulo_pagina(ws, 'Remanescentes — Juntas',
                  'Sem "Liberado END" (dt_lib_END) — %d de %d juntas | Fonte: %s, aba MAPA_JUNTA'
                  % (len(remanescentes_jt), len(validas_jt), juntas_path.split('/')[-1]),
                  len(COLUNAS_JUNTAS))
    escrever_tabela_bruta(ws, COLUNAS_JUNTAS, remanescentes_jt, 4, 'TabRemanescentesJuntas')
    ws.sheet_view.showGridLines = False

    # -------------------------------------------------- Acompanhamento -----
    ws = wb.create_sheet('Acompanhamento_Fab_Montagem')
    titulo_pagina(ws, 'Acompanhamento de Fabricação e Montagem de Tubulação',
                  'Escopo novo (remanescente) por unidade e pacote SOP, e cronograma mensal a partir de %s'
                  % data_inicio.strftime('%m/%Y'), 8)

    ws.cell(4, 1, 'Escopo novo (peso remanescente, ton)').font = FONTE_ROTULO
    CELULA_ESCOPO = 'B4'
    c = ws.cell(4, 2, peso_remanescente)
    c.number_format, c.font = FMT_TON, Font(bold=True, color=AZUL)
    ws.cell(5, 1, 'Data de início da retomada').font = FONTE_ROTULO
    ws.cell(5, 2, data_inicio).number_format = FMT_DATA
    ws.cell(6, 1, 'Prazo contratual (meses)').font = FONTE_ROTULO
    ws.cell(6, 2, prazo_meses).number_format = FMT_INT
    ws.cell(7, 1, 'Término previsto').font = FONTE_ROTULO
    ws.cell(7, 2, data_fim).number_format = FMT_DATA
    for lin in (4, 5, 6, 7):
        ws.cell(lin, 1).fill = PREENCH_ROTULO
        ws.cell(lin, 1).border = ws.cell(lin, 2).border = BORDA_FINA

    linha = 10
    ws.cell(linha, 1, 'Escopo novo por Unidade').font = Font(bold=True, size=12, color=AZUL)
    linha += 1
    cab_un = ['Unidade', 'Spools Remanescentes', 'Spools Total', 'Peso Remanescente (ton)', 'Peso Total (ton)', '% Remanescente']
    for j, nome in enumerate(cab_un, start=1):
        c = ws.cell(linha, j, nome)
        c.font, c.fill = FONTE_CABECALHO, PREENCH_CABECALHO
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    linha_cab_un = linha
    linha += 1
    for unidade in sorted(por_unidade, key=lambda u: -por_unidade[u]['peso_remanescente']):
        a = por_unidade[unidade]
        ws.cell(linha, 1, unidade)
        ws.cell(linha, 2, a['spools_remanescente']).number_format = FMT_INT
        ws.cell(linha, 3, a['spools_total']).number_format = FMT_INT
        ws.cell(linha, 4, round(a['peso_remanescente'] / 1000, 3)).number_format = FMT_TON
        ws.cell(linha, 5, round(a['peso_total'] / 1000, 3)).number_format = FMT_TON
        ws.cell(linha, 6, (a['peso_remanescente'] / a['peso_total']) if a['peso_total'] else 0).number_format = FMT_PCT
        linha += 1
    tabela = Table(displayName='TabResumoUnidade', ref='A%d:F%d' % (linha_cab_un, linha - 1))
    tabela.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    ws.add_table(tabela)

    linha += 2
    ws.cell(linha, 1, 'Escopo novo por Pacote SOP (30 maiores)').font = Font(bold=True, size=12, color=AZUL)
    linha += 1
    cab_sop = ['Pacote SOP', 'Linhas', 'Spools Remanescentes', 'Spools Total', 'Peso Remanescente (ton)', 'Peso Total (ton)', '% Remanescente']
    for j, nome in enumerate(cab_sop, start=1):
        c = ws.cell(linha, j, nome)
        c.font, c.fill = FONTE_CABECALHO, PREENCH_CABECALHO
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    linha_cab_sop = linha
    linha += 1
    sops_ordenados = sorted(por_sop, key=lambda s: -por_sop[s]['peso_remanescente'])[:30]
    for sop in sops_ordenados:
        a = por_sop[sop]
        ws.cell(linha, 1, sop)
        ws.cell(linha, 2, a['linhas']).number_format = FMT_INT
        ws.cell(linha, 3, a['spools_remanescente']).number_format = FMT_INT
        ws.cell(linha, 4, a['spools_total']).number_format = FMT_INT
        ws.cell(linha, 5, round(a['peso_remanescente'] / 1000, 3)).number_format = FMT_TON
        ws.cell(linha, 6, round(a['peso_total'] / 1000, 3)).number_format = FMT_TON
        ws.cell(linha, 7, (a['peso_remanescente'] / a['peso_total']) if a['peso_total'] else 0).number_format = FMT_PCT
        linha += 1
    tabela = Table(displayName='TabResumoSOP', ref='A%d:G%d' % (linha_cab_sop, linha - 1))
    tabela.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    ws.add_table(tabela)

    linha += 2
    ws.cell(linha, 1, 'Cronograma Mensal (%d meses — retomada)' % prazo_meses).font = Font(bold=True, size=12, color=AZUL)
    linha += 1
    ws.cell(linha, 1, 'Peso Planejado = sugestão de curva linear (escopo novo ÷ %d meses) — edite para refletir o planejamento real.' % prazo_meses).font = FONTE_SUBTITULO
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=8)
    linha += 1
    ws.cell(linha, 1, 'Peso Realizado fica em branco — preencha mês a mês conforme a obra avançar.').font = FONTE_SUBTITULO
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=8)
    linha += 1
    cab_cron = ['Mês', 'Fim do Mês', 'Nº do Mês', 'Peso Planejado Acumulado (ton)',
                'Peso Realizado Acumulado (ton)', '% Avanço Realizado', 'Saldo a Realizar (ton)']
    for j, nome in enumerate(cab_cron, start=1):
        c = ws.cell(linha, j, nome)
        c.font, c.fill = FONTE_CABECALHO, PREENCH_CABECALHO
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    linha_cab_cron = linha
    primeira_linha_dados = linha + 1
    linha += 1
    for i in range(prazo_meses):
        mes_idx = (data_inicio.month - 1 + i) % 12
        ano = data_inicio.year + (data_inicio.month - 1 + i) // 12
        fim_do_mes = datetime.date(ano, mes_idx + 1, calendar.monthrange(ano, mes_idx + 1)[1])
        ws.cell(linha, 1, '%s/%s' % (MESES_ABREV[mes_idx], str(ano)[2:]))
        ws.cell(linha, 2, fim_do_mes).number_format = FMT_DATA
        ws.cell(linha, 3, i + 1).number_format = FMT_INT
        ws.cell(linha, 4, round(peso_remanescente * (i + 1) / prazo_meses, 3)).number_format = FMT_TON
        ws.cell(linha, 5, None).number_format = FMT_TON
        ws.cell(linha, 6, '=E%d/$%s' % (linha, CELULA_ESCOPO)).number_format = FMT_PCT
        ws.cell(linha, 7, '=$%s-E%d' % (CELULA_ESCOPO, linha)).number_format = FMT_TON
        linha += 1
    fim_cron = linha - 1
    tabela = Table(displayName='TabCronograma', ref='A%d:G%d' % (linha_cab_cron, fim_cron))
    tabela.tableStyleInfo = TableStyleInfo(name='TableStyleMedium9', showRowStripes=True)
    ws.add_table(tabela)

    grafico = LineChart()
    grafico.title = 'Curva de avanço — planejado x realizado (ton)'
    grafico.y_axis.title, grafico.x_axis.title = 'ton', 'Mês'
    grafico.add_data(Reference(ws, min_col=4, max_col=5, min_row=linha_cab_cron, max_row=fim_cron), titles_from_data=True)
    grafico.set_categories(Reference(ws, min_col=1, min_row=primeira_linha_dados, max_row=fim_cron))
    grafico.height, grafico.width = 10, 22
    ws.add_chart(grafico, 'I10')

    ws.column_dimensions['A'].width = 30
    for col, w in zip('BCDEFG', (14, 24, 24, 24, 18, 18)):
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A%d' % primeira_linha_dados

    wb._sheets = [wb[n] for n in ('Leia-me', 'Resumo', 'Acompanhamento_Fab_Montagem',
                                  'Remanescentes_Spools', 'Remanescentes_Juntas')]
    wb.active = 0
    wb.save(saida)

    print('Spools ......... %d válidos, %d remanescentes (%.3f ton), %d concluídos' % (
        len(validas_sp), len(remanescentes_sp), peso_remanescente, concluidos_sp_n))
    print('Juntas ......... %d válidas, %d remanescentes, %d concluídas' % (
        len(validas_jt), len(remanescentes_jt), concluidas_jt_n))
    print('Peso total ..... %.3f ton (%.1f%% remanescente)' % (peso_total, pct_remanescente * 100))
    print('Cronograma ..... %s a %s (%d meses)' % (
        data_inicio.strftime('%m/%Y'), data_fim.strftime('%m/%Y'), prazo_meses))
    print('Arquivo salvo em', saida)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--spools', required=True, help='Arquivo original do ControlTub, aba "SGS"')
    ap.add_argument('--juntas', required=True, help='Arquivo original do ControlTub, aba "MAPA_JUNTA"')
    ap.add_argument('--saida', default='retomada/ControlTub-Retomada-Out2026.xlsx')
    ap.add_argument('--data-inicio', default='2026-10-01', help='YYYY-MM-DD')
    ap.add_argument('--prazo-meses', type=int, default=32)
    args = ap.parse_args()
    data_inicio = datetime.date(*(int(x) for x in args.data_inicio.split('-')))
    montar(args.spools, args.juntas, args.saida, data_inicio, args.prazo_meses)


if __name__ == '__main__':
    main()
